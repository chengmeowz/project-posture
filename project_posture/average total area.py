# -*- coding: utf-8 -*-
# author: Sunny Qu
# modified by: Cheng Zheng (add path, so the address don't need to change for different users.)

import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook

path_current = os.getcwd() # get current path, i.e this python file directory
path_project = os.path.dirname(path_current) # go up one directory; find the project directory

experiment = pd.read_excel(path_project + '/BDS/BDSinfo_modified.xlsx') # read first sheet from xlsx

column = ['Total Area (cm²)', 'MVPA_minutes.week', 'Subject', 'Vision', 'Surface']
column_index = [list(experiment.columns).index(i) for i in column]
exp_ = np.array(experiment.iloc[:,column_index]).tolist()

ave_dict = dict()
ave_num = dict() # create empty dict
for i in exp_:
    if tuple(i[1:]) in ave_dict.keys():
        ave_dict[tuple(i[1:])] = ave_dict[tuple(i[1:])] + i[0]
        ave_num[tuple(i[1:])] += 1
    else:
        ave_dict[tuple(i[1:])] = i[0]
        ave_num[tuple(i[1:])] = 1

wb = load_workbook(path_project + '/BDS/aveTotalArea_MVPA.xlsx')
ws = wb.create_sheet(index = 0, title = 'Average Total Area')

ws['A1'] = 'aveTotalArea'
ws['B1'] = column[1]
ws['C1'] = column[2]
ws['D1'] = column[3]
ws['E1'] = column[4]

for i in ave_dict.keys():
    outline = [ave_dict[i]/ave_num[i]] + list(i)
    ws.append(outline)
wb.save(path_project + '/BDS/aveTotalArea_MVPA.xlsx')