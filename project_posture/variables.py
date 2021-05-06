# -*- coding: utf-8 -*-
# author: Sunny Qu
# modified by: Cheng Zheng (add path, so the address don't need to change for different users.)

import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook

# Get the project directory
path_current = os.getcwd() # get current path, aka this python file directory
path_project = os.path.dirname(path_current) # go up one directory, aka find the project directory

experiment = pd.read_excel(path_project + '/BDS/BDSinfo_modified.xlsx') # read first sheet from xlsx
column = ['Total Area (cm²)','AP RMS (cm)','ML RMS (cm)','Total Displacement (cm)','Total Velocity (cm/s)', 'MVPA_minutes.week', 'Subject', 'Vision', 'Surface']
column_index = [list(experiment.columns).index(i) for i in column]
exp_ = np.array(experiment.iloc[:,column_index]).tolist()

ave_dict_area = dict()
ave_num_area = dict() # create empty dict
for i in exp_:
    if tuple(i[5:]) in ave_dict_area.keys():
        ave_dict_area[tuple(i[5:])] = ave_dict_area[tuple(i[5:])] + i[0]
        ave_num_area[tuple(i[5:])] += 1
    else:
        ave_dict_area[tuple(i[5:])] = i[0]
        ave_num_area[tuple(i[5:])] = 1

ave_dict_AP_RMS = dict()
ave_num_AP_RMS = dict() # create empty dict
for i in exp_:
    if tuple(i[5:]) in ave_dict_AP_RMS.keys():
        ave_dict_AP_RMS[tuple(i[5:])] = ave_dict_AP_RMS[tuple(i[5:])] + i[1]
        ave_num_AP_RMS[tuple(i[5:])] += 1
    else:
        ave_dict_AP_RMS[tuple(i[5:])] = i[1]
        ave_num_AP_RMS[tuple(i[5:])] = 1


ave_dict_ML_MRS = dict()
ave_num_ML_RMS = dict() # create empty dict
for i in exp_:
    if tuple(i[5:]) in ave_dict_ML_MRS.keys():
        ave_dict_ML_MRS[tuple(i[5:])] = ave_dict_ML_MRS[tuple(i[5:])] + i[2]
        ave_num_ML_RMS[tuple(i[5:])] += 1
    else:
        ave_dict_ML_MRS[tuple(i[5:])] = i[2]
        ave_num_ML_RMS[tuple(i[5:])] = 1


ave_dict_displacement = dict()
ave_num_displacement = dict() # create empty dict
for i in exp_:
    if tuple(i[5:]) in ave_dict_displacement.keys():
        ave_dict_displacement[tuple(i[5:])] = ave_dict_displacement[tuple(i[5:])] + i[3]
        ave_num_displacement[tuple(i[5:])] += 1
    else:
        ave_dict_displacement[tuple(i[5:])] = i[3]
        ave_num_displacement[tuple(i[5:])] = 1

ave_dict_velocity = dict()
ave_num_velocity = dict() # create empty dict
for i in exp_:
    if tuple(i[5:]) in ave_dict_velocity.keys():
        ave_dict_velocity[tuple(i[5:])] = ave_dict_velocity[tuple(i[5:])] + i[4]
        ave_num_velocity[tuple(i[5:])] += 1
    else:
        ave_dict_velocity[tuple(i[5:])] = i[4]
        ave_num_velocity[tuple(i[5:])] = 1


wb = load_workbook(path_project + '/BDS/variables.xlsx')
ws = wb.create_sheet(index = 0, title = 'Average Total Area')

ws['A1'] = 'aveTotalArea'
ws['B1'] = 'aveAP_RMS'
ws['C1'] = 'aveML_RMS'
ws['D1'] = 'aveDisplacement'
ws['E1'] = 'aveVelocity'
ws['F1'] = column[5]
ws['G1'] = column[6]
ws['H1'] = column[7]
ws['I1'] = column[8]

for i in ave_dict_area.keys():
    outline = [ave_dict_area[i]/ave_num_area[i]] + [ave_dict_AP_RMS[i]/ave_num_AP_RMS[i]]+ [ave_dict_ML_MRS[i]/ave_num_ML_RMS[i]]+[ave_dict_displacement[i]/ave_num_displacement[i]]+[ave_dict_velocity[i]/ave_num_velocity[i]]+list(i)
    ws.append(outline)


wb.save(path_project + '/BDS/variables.xlsx')

